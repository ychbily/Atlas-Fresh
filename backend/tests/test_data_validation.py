"""
Automated unit tests for Excel dataset loading, API endpoints, and server-side validation rules.
"""

from pathlib import Path
import pytest
import openpyxl
from fastapi.testclient import TestClient

from app.main import app
from app.data_loader import load_dataset, DataValidationError, DEFAULT_DATA_PATH

client = TestClient(app)


def safe_load_dataset():
    """Safely load dataset or fail with a clear assertion message if Excel validation fails."""
    try:
        return load_dataset()
    except DataValidationError as exc:
        err_msgs = [f"{e.sheet}/{e.entity_id or 'General'}: {e.message}" for e in exc.errors]
        raise pytest.fail(f"Excel dataset validation failed: {err_msgs}") from None


def test_dataset_loading_and_api_contract():
    """Verify Excel loading, data aggregation, structural invariants, and GET /api/data endpoint."""
    dataset = safe_load_dataset()

    assert len(dataset.farms) == 20
    assert len(dataset.clients) == 10
    assert dataset.station.export_conditioning_capacity_t > 0

    exp_sum = sum(f.expected_daily_capacity_t for f in dataset.farms)
    act_a = sum(f.actual_A_t for f in dataset.farms)
    act_b = sum(f.actual_B_t for f in dataset.farms)
    act_c = sum(f.actual_C_t for f in dataset.farms)
    act_d = sum(f.actual_D_t for f in dataset.farms)
    act_tot = act_a + act_b + act_c + act_d

    assert dataset.total_expected_capacity_t == exp_sum
    assert dataset.total_actual_supply_t == act_tot
    assert dataset.actual_by_segment_t == {"A": act_a, "B": act_b, "C": act_c, "D": act_d}

    # API Endpoint check
    resp = client.get("/api/data")
    assert resp.status_code == 200, f"GET /api/data failed ({resp.status_code}): {resp.json()}"
    body = resp.json()
    assert body["total_actual_supply_t"] == act_tot


def test_server_validation_rules(tmp_path: Path):
    """Verify server-side rejection of invalid mix sums, acceptance modes, non-5t multiples, and negative values."""
    # 1. Invalid mix sum (!= 1.0)
    wb1 = openpyxl.load_workbook(DEFAULT_DATA_PATH)
    wb1["Farms"].cell(5, 4).value = 0.95
    bad_mix = tmp_path / "bad_mix.xlsx"
    wb1.save(bad_mix)
    with pytest.raises(DataValidationError) as exc1:
        load_dataset(bad_mix)
    assert any("Mix percentages sum to" in err.message for err in exc1.value.errors)

    # 2. Invalid acceptance mode
    wb2 = openpyxl.load_workbook(DEFAULT_DATA_PATH)
    wb2["Clients"].cell(5, 3).value = "INVALID_MODE"
    bad_mode = tmp_path / "bad_mode.xlsx"
    wb2.save(bad_mode)
    with pytest.raises(DataValidationError) as exc2:
        load_dataset(bad_mode)
    assert any("Acceptance mode must be 'EXACT' or 'MINIMUM'" in err.message for err in exc2.value.errors)

    # 3. Non-5t multiple quantity
    wb3 = openpyxl.load_workbook(DEFAULT_DATA_PATH)
    wb3["Farms"].cell(5, 8).value = 23
    bad_step = tmp_path / "bad_step.xlsx"
    wb3.save(bad_step)
    with pytest.raises(DataValidationError) as exc3:
        load_dataset(bad_step)
    assert any("must be a multiple of 5" in err.message for err in exc3.value.errors)

    # 4. Negative actual quantity
    wb4 = openpyxl.load_workbook(DEFAULT_DATA_PATH)
    wb4["Farms"].cell(5, 8).value = -10
    bad_neg = tmp_path / "bad_neg.xlsx"
    wb4.save(bad_neg)
    with pytest.raises(DataValidationError) as exc4:
        load_dataset(bad_neg)
    assert any("cannot be negative" in err.message for err in exc4.value.errors)
