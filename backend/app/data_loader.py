"""
Excel data loader and validator for the Atlas Fresh planning workspace.

Reads the authoritative Excel workbook, performs server-side business validation,
and returns structured Pydantic models for farms, clients, and station.
"""

from pathlib import Path
from typing import Optional
import openpyxl

from app.models import (
    Farm,
    Client,
    Station,
    DatasetResponse,
    QualitySegment,
    AcceptanceMode,
    ValidationErrorDetail,
)

# Default location for the authoritative Excel dataset
DEFAULT_DATA_PATH = Path(__file__).resolve().parent.parent / "data" / "Atlas_Fresh_Production_Commercial_Data.xlsx"


class DataValidationError(Exception):
    """Raised when Excel workbook fails validation checks."""

    def __init__(self, errors: list[ValidationErrorDetail]) -> None:
        self.errors = errors
        super().__init__(f"Validation failed with {len(errors)} error(s).")


def _is_multiple_of_five(value: float) -> bool:
    """
    Check if a numeric value is an exact multiple of 5 tonnes.

    Args:
        value (float): Tonnes to check.

    Returns:
        bool: True if value is an exact multiple of 5, False otherwise.
    """
    return abs(value % 5) < 1e-6


def _validate_farm_row(
    f_id: str,
    exp_cap: float,
    mix: tuple[float, float, float, float],
    actuals: tuple[float, float, float, float],
    errors: list[ValidationErrorDetail],
) -> None:
    """
    Validate business rules for a single farm row.

    Args:
        f_id (str): Farm identifier.
        exp_cap (float): Planned daily capacity in tonnes.
        mix (tuple[float, float, float, float]): Expected percentages (A, B, C, D).
        actuals (tuple[float, float, float, float]): Actual tonnes (A, B, C, D).
        errors (list[ValidationErrorDetail]): Shared error list to append to.
    """
    if exp_cap <= 0:
        errors.append(ValidationErrorDetail(
            sheet="Farms", entity_id=f_id, field="expected_daily_capacity_t",
            message="Expected capacity must be greater than 0",
        ))

    mix_sum = sum(mix)
    if abs(mix_sum - 1.0) > 1e-4:
        errors.append(ValidationErrorDetail(
            sheet="Farms", entity_id=f_id, field="expected_mix_sum",
            message=f"Mix percentages sum to {mix_sum:.4f}, must equal exactly 1.0",
        ))

    for seg, pct in zip(["A", "B", "C", "D"], mix):
        if pct < 0.0 or pct > 1.0:
            errors.append(ValidationErrorDetail(
                sheet="Farms", entity_id=f_id, field=f"expected_{seg}_pct",
                message=f"Expected {seg} percentage must be between 0.0 and 1.0",
            ))

    for seg, act_t in zip(["A", "B", "C", "D"], actuals):
        if act_t < 0:
            errors.append(ValidationErrorDetail(
                sheet="Farms", entity_id=f_id, field=f"actual_{seg}_t",
                message=f"Actual {seg} tonnes cannot be negative",
            ))
        elif not _is_multiple_of_five(act_t):
            errors.append(ValidationErrorDetail(
                sheet="Farms", entity_id=f_id, field=f"actual_{seg}_t",
                message=f"Actual {seg} tonnes ({act_t}) must be a multiple of 5",
            ))


def _parse_farms(sheet: openpyxl.worksheet.worksheet.Worksheet, errors: list[ValidationErrorDetail]) -> list[Farm]:
    """
    Parse and validate rows from the 'Farms' worksheet.

    Args:
        sheet (Worksheet): The openpyxl Farms worksheet.
        errors (list[ValidationErrorDetail]): Shared list to append validation errors.

    Returns:
        list[Farm]: Validated list of Farm model instances.
    """
    farms: list[Farm] = []
    seen_ids: set[str] = set()

    for row_idx in range(5, sheet.max_row + 1):
        f_id = sheet.cell(row_idx, 1).value
        f_name = sheet.cell(row_idx, 2).value
        if f_id is None and f_name is None:
            continue

        farm_id_str = str(f_id).strip() if f_id is not None else ""
        if not farm_id_str:
            errors.append(ValidationErrorDetail(
                sheet="Farms", entity_id=f"Row {row_idx}", field="farm_id",
                message="Farm ID cannot be empty",
            ))
            continue

        if farm_id_str in seen_ids:
            errors.append(ValidationErrorDetail(
                sheet="Farms", entity_id=farm_id_str, field="farm_id",
                message=f"Duplicate farm ID: {farm_id_str}",
            ))
        seen_ids.add(farm_id_str)

        try:
            exp_cap = float(sheet.cell(row_idx, 3).value or 0.0)
            exp_mix = (
                float(sheet.cell(row_idx, 4).value or 0.0),
                float(sheet.cell(row_idx, 5).value or 0.0),
                float(sheet.cell(row_idx, 6).value or 0.0),
                float(sheet.cell(row_idx, 7).value or 0.0),
            )
            act_t = (
                float(sheet.cell(row_idx, 8).value or 0.0),
                float(sheet.cell(row_idx, 9).value or 0.0),
                float(sheet.cell(row_idx, 10).value or 0.0),
                float(sheet.cell(row_idx, 11).value or 0.0),
            )
        except (ValueError, TypeError) as exc:
            errors.append(ValidationErrorDetail(
                sheet="Farms", entity_id=farm_id_str, field="numeric_values",
                message=f"Non-numeric value in row: {exc}",
            ))
            continue

        _validate_farm_row(farm_id_str, exp_cap, exp_mix, act_t, errors)

        farms.append(Farm(
            farm_id=farm_id_str,
            farm_name=str(f_name).strip() if f_name is not None else f"Farm {farm_id_str}",
            expected_daily_capacity_t=exp_cap,
            expected_A_pct=exp_mix[0],
            expected_B_pct=exp_mix[1],
            expected_C_pct=exp_mix[2],
            expected_D_pct=exp_mix[3],
            actual_A_t=act_t[0],
            actual_B_t=act_t[1],
            actual_C_t=act_t[2],
            actual_D_t=act_t[3],
        ))

    return farms


def _validate_client_row(
    c_id: str,
    mode_str: str,
    seg_str: str,
    demand_t: float,
    price_eur: float,
    errors: list[ValidationErrorDetail],
) -> None:
    """
    Validate commercial constraints for a single client row.

    Args:
        c_id (str): Client ID.
        mode_str (str): Acceptance mode string.
        seg_str (str): Requested segment string.
        demand_t (float): Demand in tonnes.
        price_eur (float): Export price in EUR/t.
        errors (list[ValidationErrorDetail]): Shared error list to append to.
    """
    if mode_str not in AcceptanceMode.__members__:
        errors.append(ValidationErrorDetail(
            sheet="Clients", entity_id=c_id, field="acceptance_mode",
            message=f"Acceptance mode must be 'EXACT' or 'MINIMUM' (got '{mode_str}')",
        ))

    if seg_str not in QualitySegment.__members__:
        errors.append(ValidationErrorDetail(
            sheet="Clients", entity_id=c_id, field="requested_segment",
            message=f"Requested segment must be 'A', 'B', 'C', or 'D' (got '{seg_str}')",
        ))

    if demand_t <= 0:
        errors.append(ValidationErrorDetail(
            sheet="Clients", entity_id=c_id, field="demand_t",
            message="Demand must be greater than 0",
        ))
    elif not _is_multiple_of_five(demand_t):
        errors.append(ValidationErrorDetail(
            sheet="Clients", entity_id=c_id, field="demand_t",
            message=f"Demand tonnes ({demand_t}) must be a multiple of 5",
        ))

    if price_eur <= 0:
        errors.append(ValidationErrorDetail(
            sheet="Clients", entity_id=c_id, field="export_price_per_t_eur",
            message="Export price must be greater than 0 EUR",
        ))


def _parse_clients(sheet: openpyxl.worksheet.worksheet.Worksheet, errors: list[ValidationErrorDetail]) -> list[Client]:
    """
    Parse and validate rows from the 'Clients' worksheet.

    Args:
        sheet (Worksheet): The openpyxl Clients worksheet.
        errors (list[ValidationErrorDetail]): Shared list to append validation errors.

    Returns:
        list[Client]: Validated list of Client model instances.
    """
    clients: list[Client] = []
    seen_ids: set[str] = set()

    for row_idx in range(5, sheet.max_row + 1):
        c_id = sheet.cell(row_idx, 1).value
        c_name = sheet.cell(row_idx, 2).value
        if c_id is None and c_name is None:
            continue

        client_id_str = str(c_id).strip() if c_id is not None else ""
        if not client_id_str:
            errors.append(ValidationErrorDetail(
                sheet="Clients", entity_id=f"Row {row_idx}", field="client_id",
                message="Client ID cannot be empty",
            ))
            continue

        if client_id_str in seen_ids:
            errors.append(ValidationErrorDetail(
                sheet="Clients", entity_id=client_id_str, field="client_id",
                message=f"Duplicate client ID: {client_id_str}",
            ))
        seen_ids.add(client_id_str)

        mode_str = str(sheet.cell(row_idx, 3).value or "").strip().upper()
        seg_str = str(sheet.cell(row_idx, 4).value or "").strip().upper()

        try:
            demand_t = float(sheet.cell(row_idx, 5).value or 0.0)
            price_eur = float(sheet.cell(row_idx, 6).value or 0.0)
        except (ValueError, TypeError) as exc:
            errors.append(ValidationErrorDetail(
                sheet="Clients", entity_id=client_id_str, field="demand_or_price",
                message=f"Non-numeric demand or price: {exc}",
            ))
            continue

        _validate_client_row(client_id_str, mode_str, seg_str, demand_t, price_eur, errors)

        safe_mode = AcceptanceMode(mode_str) if mode_str in AcceptanceMode.__members__ else AcceptanceMode.EXACT
        safe_seg = QualitySegment(seg_str) if seg_str in QualitySegment.__members__ else QualitySegment.A

        clients.append(Client(
            client_id=client_id_str,
            client_name=str(c_name).strip() if c_name is not None else f"Client {client_id_str}",
            acceptance_mode=safe_mode,
            requested_segment=safe_seg,
            demand_t=demand_t,
            export_price_per_t_eur=price_eur,
        ))

    return clients


def _parse_station(sheet: openpyxl.worksheet.worksheet.Worksheet, errors: list[ValidationErrorDetail]) -> Station:
    """
    Parse and validate configuration from the 'Station' worksheet.

    Args:
        sheet (Worksheet): The openpyxl Station worksheet.
        errors (list[ValidationErrorDetail]): Shared list to append validation errors.

    Returns:
        Station: Validated Station model instance.
    """
    station_id = str(sheet.cell(5, 1).value or "STATION-01").strip()

    try:
        capacity_t = float(sheet.cell(5, 2).value or 500.0)
        local_ratio = float(sheet.cell(5, 3).value or 0.10)
    except (ValueError, TypeError) as exc:
        errors.append(ValidationErrorDetail(
            sheet="Station", entity_id=station_id, field="capacity_or_ratio",
            message=f"Non-numeric capacity or local ratio: {exc}",
        ))
        capacity_t, local_ratio = 500.0, 0.10

    if capacity_t <= 0 or not _is_multiple_of_five(capacity_t):
        errors.append(ValidationErrorDetail(
            sheet="Station", entity_id=station_id, field="export_conditioning_capacity_t",
            message=f"Station capacity ({capacity_t}) must be positive and a multiple of 5",
        ))

    if local_ratio <= 0 or local_ratio > 1.0:
        errors.append(ValidationErrorDetail(
            sheet="Station", entity_id=station_id, field="local_market_ratio",
            message=f"Local ratio ({local_ratio}) must be between 0.0 and 1.0",
        ))

    ref_prices: dict[str, float] = {}
    for row_idx in range(17, 21):
        seg = str(sheet.cell(row_idx, 1).value or "").strip().upper()
        if not seg:
            continue
        try:
            p_val = float(sheet.cell(row_idx, 2).value or 0.0)
            if p_val <= 0:
                errors.append(ValidationErrorDetail(
                    sheet="Station", entity_id=station_id, field=f"ref_price_{seg}",
                    message=f"Reference price for segment {seg} must be positive",
                ))
            ref_prices[seg] = p_val
        except (ValueError, TypeError) as exc:
            errors.append(ValidationErrorDetail(
                sheet="Station", entity_id=station_id, field=f"ref_price_{seg}",
                message=f"Invalid reference price for segment {seg}: {exc}",
            ))

    for required_seg in ["A", "B", "C", "D"]:
        if required_seg not in ref_prices:
            errors.append(ValidationErrorDetail(
                sheet="Station", entity_id=station_id, field="reference_prices",
                message=f"Missing reference price for segment {required_seg}",
            ))

    return Station(
        station_id=station_id,
        export_conditioning_capacity_t=capacity_t,
        local_market_ratio=local_ratio,
        reference_prices=ref_prices,
    )


def load_dataset(file_path: Optional[str | Path] = None) -> DatasetResponse:
    """
    Load, parse, and validate the Atlas Fresh production and commercial dataset.

    Args:
        file_path (Optional[str | Path]): Path to Excel workbook. Uses default data file if None.

    Returns:
        DatasetResponse: Aggregated dataset containing validated farms, clients, and station.

    Raises:
        FileNotFoundError: If the specified Excel file does not exist.
        DataValidationError: If any sheet data violates business rules.
    """
    path = Path(file_path) if file_path else DEFAULT_DATA_PATH

    if not path.exists():
        raise FileNotFoundError(f"Authoritative dataset file not found at: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    required_sheets = ["Farms", "Clients", "Station"]
    missing_sheets = [s for s in required_sheets if s not in wb.sheetnames]

    if missing_sheets:
        raise DataValidationError([
            ValidationErrorDetail(
                sheet="Workbook",
                message=f"Missing required sheet(s): {', '.join(missing_sheets)}",
            )
        ])

    errors: list[ValidationErrorDetail] = []
    farms = _parse_farms(wb["Farms"], errors)
    clients = _parse_clients(wb["Clients"], errors)
    station = _parse_station(wb["Station"], errors)

    if errors:
        raise DataValidationError(errors)

    total_exp = sum(f.expected_daily_capacity_t for f in farms)
    act_a = sum(f.actual_A_t for f in farms)
    act_b = sum(f.actual_B_t for f in farms)
    act_c = sum(f.actual_C_t for f in farms)
    act_d = sum(f.actual_D_t for f in farms)
    total_act = act_a + act_b + act_c + act_d

    return DatasetResponse(
        farms=farms,
        clients=clients,
        station=station,
        total_expected_capacity_t=total_exp,
        total_actual_supply_t=total_act,
        actual_by_segment_t={
            "A": act_a,
            "B": act_b,
            "C": act_c,
            "D": act_d,
        },
    )
